import os
import re
from openpyxl import Workbook
wb = Workbook()
ws = wb.active


work_file = os.path.basename(r'C:\Users\Asus\PycharmProjects\nu-analysis2\data.txt')
if os.path.isfile(work_file):
    print('Рабочий файл: ' + work_file)


file = open(work_file, 'r')

try:
    txt = file.read()
finally:
    file.close()


p = re.compile("(\w+)")
res = p.findall(txt)
res2 = []
res3 = []

for k in range(len(res)-1):
    res2.append(res[k]+' '+res[k+1])

for k in range(len(res)-2):
    res3.append(res[k]+' '+res[k+1]+' '+res[k+2])

lsWord = {}
for key in res:
    key = key.lower()
    if key in lsWord:
        value = lsWord[key]
        lsWord[key] = value + 1
    else:
        lsWord[key] = 1

for key in res2:
    key = key.lower()
    if key in lsWord:
        value = lsWord[key]
        lsWord[key] = value + 1
    else:
        lsWord[key] = 1

for key in res3:
    key = key.lower()
    if key in lsWord:
        value = lsWord[key]
        lsWord[key] = value + 1
    else:
        lsWord[key] = 1

c = 1
for i in lsWord:
    ws['A'+str(c)] = int(lsWord[i])
    ws['B'+str(c)] = str(i)
    c = c+1

wb.save("save.xlsx")
