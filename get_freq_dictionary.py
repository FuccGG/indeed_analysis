# -*- coding: utf-8 -*-
import os
import re
import sys
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

#ССылка на файл с текстом
work_file = os.path.basename(r'data.txt')
if os.path.isfile(work_file):
    print('Рабочий файл: ' + work_file)

# проблемы с декодированием в utf-8, необработанные символы игнорируются
file = open(work_file, encoding='utf-8', errors='ignore')

try:
    txt = file.read()
finally:
    file.close()


p = re.compile("[a-zA-Zа-яА-Я0-9\+\#]+\.{0,1}[a-zA-Zа-яА-Я0-9\+\#]+")
res = p.findall(txt)
res2 = []
res3 = []
resClear = []

# В этот список добавлять слова, которые подлежат дезинтеграции
resNor = ['and', 'to', 'the', 'of', 'in', 'with', 'for', 'or', 'is', 'as', 'you', 'on', 'be', 'are', 'our', 'will',
          'we', 'an', 'that', 'at', 'the', 'фактически', 'получается']

for k in res:
    k = k.lower()
    if k not in resNor:
        resClear.append(k)


for k in range(len(resClear) - 1):
    res2.append(resClear[k] + ' ' + resClear[k + 1])

for k in range(len(resClear) - 2):
    res3.append(resClear[k] + ' ' + resClear[k + 1] + ' ' + resClear[k + 2])



lsWord = {}
for key in resClear:
    key = key.lower()
    if key in lsWord:
        value = lsWord[key]
        lsWord[key] = value + 1
    else:
        lsWord[key] = 1

for key in res2:
    key = key.lower()
    if key in lsWord:
        value = lsWord[key]
        lsWord[key] = value + 1
    else:
        lsWord[key] = 1

for key in res3:
    key = key.lower()
    if key in lsWord:
        value = lsWord[key]
        lsWord[key] = value + 1
    else:
        lsWord[key] = 1

c = 1
for i in lsWord:
    ws['A'+str(c)] = int(lsWord[i])
    ws['B'+str(c)] = str(i)
    c = c+1


if os.path.exists("save.xlsx"):
    os.remove("save.xlsx")

wb.save("save.xlsx")


